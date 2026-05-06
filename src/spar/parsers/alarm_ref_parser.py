"""Samsung RAN 알람 레퍼런스 Excel 파서.

Excel 구조:
    - Sheet 0: 가이드 시트 (스킵)
    - Sheet 1+: 알람 목록 시트 (파싱 대상)

알람 코드 형식:
    - RAN:     A + 7자리 숫자 + R  (총 9자, 예: A1234567R)
    - Non-RAN: A 또는 H + 3자리 숫자 + D  (총 5자, 예: A123D, H456D)

지원 컬럼 (헤더명 자동 탐색, 별칭 허용):
    Alarm Code, Alarm Name, Severity, Group
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

import openpyxl

_RAN_CODE_RE = re.compile(r'^A\d{7}R$')
_NON_RAN_CODE_RE = re.compile(r'^[AH]\d{3}D$')

_COLUMN_ALIASES: dict[str, str] = {
    "alarm code": "alarm_id",
    "alarm id": "alarm_id",
    "code": "alarm_id",
    "id": "alarm_id",
    "alarm name": "alarm_name",
    "name": "alarm_name",
    "severity": "severity",
    "level": "severity",
    "group": "group",
    "category": "group",
}

REQUIRED_FIELDS = {"alarm_id", "alarm_name"}

AlarmType = Literal["RAN", "NON_RAN", "UNKNOWN"]


def _normalize_alarm_id(raw: str) -> str:
    return raw.strip().upper()


def _classify_alarm_code(code: str) -> AlarmType:
    if _RAN_CODE_RE.match(code):
        return "RAN"
    if _NON_RAN_CODE_RE.match(code):
        return "NON_RAN"
    return "UNKNOWN"


@dataclass
class AlarmRecord:
    alarm_id: str
    alarm_name: str
    severity: str = ""
    group: str = ""
    alarm_type: AlarmType = "UNKNOWN"

    def __post_init__(self) -> None:
        self.alarm_id = _normalize_alarm_id(self.alarm_id)
        if self.alarm_type == "UNKNOWN":
            self.alarm_type = _classify_alarm_code(self.alarm_id)

    def to_chunk_text(self) -> str:
        lines = [f"Alarm: {self.alarm_id} — {self.alarm_name}"]
        if self.alarm_type != "UNKNOWN":
            lines.append(f"Type: {self.alarm_type}")
        if self.severity:
            lines.append(f"Severity: {self.severity}")
        if self.group:
            lines.append(f"Group: {self.group}")
        return "\n".join(lines)

    def to_keywords(self) -> list[str]:
        return [v for v in (self.alarm_id, self.alarm_name, self.severity,
                            self.group, self.alarm_type) if v]

    def to_dict(self) -> dict[str, Any]:
        return {
            "alarm_id": self.alarm_id,
            "alarm_name": self.alarm_name,
            "severity": self.severity,
            "group": self.group,
            "alarm_type": self.alarm_type,
        }


@dataclass
class AlarmRefParseResult:
    records: list[AlarmRecord] = field(default_factory=list)
    skipped_rows: int = 0
    warnings: list[str] = field(default_factory=list)


def _resolve_header_row(ws) -> tuple[int, dict[str, int]]:
    for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            alias = str(cell).strip().lower()
            field_name = _COLUMN_ALIASES.get(alias)
            if field_name and field_name not in col_map:
                col_map[field_name] = col_idx
        if REQUIRED_FIELDS.issubset(col_map):
            return row_idx, col_map
    raise ValueError(
        f"헤더 행 탐색 실패 — 필수 컬럼 없음: {REQUIRED_FIELDS}. "
        "컬럼명 확인 필요 (Alarm Code, Alarm Name)"
    )


def _cell_str(row: tuple, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    return str(val).strip()


def _parse_sheet(ws, result: AlarmRefParseResult, sheet_name: str) -> None:
    try:
        header_row, col_map = _resolve_header_row(ws)
    except ValueError as e:
        result.warnings.append(f"시트 '{sheet_name}' 스킵: {e}")
        return

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if all(v is None for v in row):
            continue

        alarm_id = _cell_str(row, col_map.get("alarm_id"))
        alarm_name = _cell_str(row, col_map.get("alarm_name"))

        if not alarm_id or not alarm_name:
            result.skipped_rows += 1
            result.warnings.append(
                f"시트 '{sheet_name}' 행 {row_idx}: alarm_id 또는 alarm_name 비어있음 — 스킵"
            )
            continue

        normalized_id = _normalize_alarm_id(alarm_id)
        alarm_type = _classify_alarm_code(normalized_id)
        if alarm_type == "UNKNOWN":
            result.warnings.append(
                f"시트 '{sheet_name}' 행 {row_idx}: 알 수 없는 알람 코드 형식 '{normalized_id}'"
            )

        result.records.append(
            AlarmRecord(
                alarm_id=alarm_id,
                alarm_name=alarm_name,
                severity=_cell_str(row, col_map.get("severity")),
                group=_cell_str(row, col_map.get("group")),
                alarm_type=alarm_type,
            )
        )


def parse_alarm_ref_excel(
    path: str | Path,
    sheet_name: str | None = None,
) -> AlarmRefParseResult:
    """Excel 파일에서 알람 목록 파싱.

    sheet_name 미지정 시: 첫 번째 시트(가이드)를 스킵하고 나머지 모든 시트 파싱.
    sheet_name 지정 시: 해당 시트만 파싱.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"파일 없음: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = AlarmRefParseResult()

    if sheet_name:
        ws = wb[sheet_name]
        _parse_sheet(ws, result, sheet_name)
    else:
        sheet_names = wb.sheetnames
        if len(sheet_names) < 2:
            result.warnings.append(
                f"시트가 1개뿐 — 가이드 시트 스킵 없이 첫 번째 시트 파싱: '{sheet_names[0]}'"
            )
            _parse_sheet(wb[sheet_names[0]], result, sheet_names[0])
        else:
            for name in sheet_names[1:]:
                _parse_sheet(wb[name], result, name)

    wb.close()
    return result
