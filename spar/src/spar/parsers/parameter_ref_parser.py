"""Samsung RAN 파라미터 레퍼런스 Excel 파서.

지원 컬럼 (순서 무관, 헤더명으로 자동 탐색):
    Hierarchy, Parameter Description, Attribute Key, Leaf Status,
    Units, Type, Pattern, Range, Default Value, Bandwidth-Dependancy,
    Config value, Level, Restriction, Service Impact, Real time change,
    Reference, Mandatory, Parameter-Family, Related Feature ID, User Level
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

# 헤더 후보명 → 정규 필드명 매핑 (소문자 비교)
_COLUMN_ALIASES: dict[str, str] = {
    # Hierarchy (col 0)
    "hierarchy": "yang_path",
    "yang path": "yang_path",
    "yang": "yang_path",
    # Parameter Description (col 1)
    "parameter description": "description",
    "description": "description",
    "desc": "description",
    # Attribute Key (col 2)
    "attribute key": "param_name",
    "attributekey": "param_name",
    "parameter name": "param_name",
    "param name": "param_name",
    "parameter": "param_name",
    "name": "param_name",
    # Leaf Status (col 3)
    "leaf status": "leaf_status",
    "leafstatus": "leaf_status",
    # Units (col 4)
    "units": "units",
    "unit": "units",
    # Type (col 5)
    "type": "type",
    # Pattern (col 6)
    "pattern": "pattern",
    # Range (col 7)
    "range": "range",
    # Default Value (col 8)
    "default value": "default",
    "default": "default",
    # Bandwidth-Dependancy (col 9)
    "bandwidth-dependancy": "bandwidth_dependency",
    "bandwidth dependancy": "bandwidth_dependency",
    "bandwidth dependency": "bandwidth_dependency",
    # Config value (col 10)
    "config value": "config_value",
    "config_value": "config_value",
    # Level (col 11)
    "level": "level",
    # Restriction (col 12)
    "restriction": "restriction",
    # Service Impact (col 13)
    "service impact": "service_impact",
    "service_impact": "service_impact",
    # Real time change (col 14)
    "real time change": "realtime_change",
    "real-time change": "realtime_change",
    # Reference (col 15)
    "reference": "reference",
    # Mandatory (col 16)
    "mandatory": "mandatory",
    # Parameter-Family (col 17)
    "parameter-family": "param_family",
    "parameter family": "param_family",
    # Related Feature ID (col 18)
    "related feature id": "related_feature_id",
    "related feature": "related_feature_id",
    # User Level (col 19)
    "user level": "user_level",
    "userlevel": "user_level",
    # legacy aliases
    "feature name": "feature_name",
    "feature": "feature_name",
    "min": "min",
    "minimum": "min",
    "max": "max",
    "maximum": "max",
}

REQUIRED_FIELDS = {"param_name", "yang_path"}


@dataclass
class ParameterRecord:
    param_name: str
    yang_path: str
    description: str = ""
    leaf_status: str = ""
    units: str = ""
    type: str = ""
    pattern: str = ""
    range: str = ""
    default: str = ""
    bandwidth_dependency: str = ""
    config_value: str = ""
    level: str = ""
    restriction: str = ""
    service_impact: str = ""
    realtime_change: str = ""
    reference: str = ""
    mandatory: str = ""
    param_family: str = ""
    related_feature_id: str = ""
    user_level: str = ""
    # legacy fields
    feature_name: str = ""
    min: str = ""
    max: str = ""

    @property
    def mo_path(self) -> list[str]:
        """YANG Path를 '/' 구분 계층 리스트로 반환."""
        return [p for p in self.yang_path.split("/") if p]

    @property
    def leaf_mo(self) -> str:
        """YANG Path의 마지막 컨테이너명 (파라미터 직상위)."""
        parts = self.mo_path
        return parts[-1] if parts else ""

    def to_chunk_text(self) -> str:
        """Milvus ingest용 텍스트 표현."""
        lines = [
            f"Parameter: {self.param_name}",
            f"Hierarchy: {self.yang_path}",
        ]
        if self.description:
            lines.append(f"Description: {self.description}")
        if self.type:
            lines.append(f"Type: {self.type}")
        if self.pattern:
            lines.append(f"Pattern: {self.pattern}")
        value_parts = []
        if self.default:
            value_parts.append(f"default={self.default}")
        if self.range:
            value_parts.append(f"range={self.range}")
        if self.min:
            value_parts.append(f"min={self.min}")
        if self.max:
            value_parts.append(f"max={self.max}")
        if value_parts:
            lines.append("Value: " + ", ".join(value_parts))
        if self.units:
            lines.append(f"Units: {self.units}")
        if self.leaf_status:
            lines.append(f"Leaf Status: {self.leaf_status}")
        if self.mandatory:
            lines.append(f"Mandatory: {self.mandatory}")
        if self.param_family:
            lines.append(f"Family: {self.param_family}")
        if self.related_feature_id:
            lines.append(f"Related Feature: {self.related_feature_id}")
        if self.feature_name:
            lines.append(f"Feature: {self.feature_name}")
        return "\n".join(lines)

    def to_dict(self) -> dict[str, Any]:
        return {
            "param_name": self.param_name,
            "yang_path": self.yang_path,
            "description": self.description,
            "leaf_status": self.leaf_status,
            "units": self.units,
            "type": self.type,
            "pattern": self.pattern,
            "range": self.range,
            "default": self.default,
            "bandwidth_dependency": self.bandwidth_dependency,
            "config_value": self.config_value,
            "level": self.level,
            "restriction": self.restriction,
            "service_impact": self.service_impact,
            "realtime_change": self.realtime_change,
            "reference": self.reference,
            "mandatory": self.mandatory,
            "param_family": self.param_family,
            "related_feature_id": self.related_feature_id,
            "user_level": self.user_level,
            "feature_name": self.feature_name,
            "min": self.min,
            "max": self.max,
        }


@dataclass
class ParameterRefParseResult:
    records: list[ParameterRecord] = field(default_factory=list)
    skipped_rows: int = 0
    warnings: list[str] = field(default_factory=list)


def _resolve_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, dict[str, int]]:
    """헤더 행 번호와 {필드명: 열 인덱스(0-based)} 반환."""
    for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            alias = str(cell).strip().lower()
            field_name = _COLUMN_ALIASES.get(alias)
            if field_name:
                col_map[field_name] = col_idx
        if REQUIRED_FIELDS.issubset(col_map):
            return row_idx, col_map
    raise ValueError(
        f"헤더 행 탐색 실패 — 필수 컬럼 없음: {REQUIRED_FIELDS}. "
        "컬럼명 확인 필요 (Parameter Name, YANG Path)"
    )


def _cell_str(row: tuple, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    return str(val).strip()


def parse_parameter_ref_excel(
    path: str | Path,
    sheet_name: str | None = None,
) -> ParameterRefParseResult:
    """Excel 파라미터 레퍼런스 파일을 파싱해 ParameterRefParseResult 반환."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"파일 없음: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row, col_map = _resolve_header_row(ws)

    result = ParameterRefParseResult()
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        if all(v is None for v in row):
            continue

        param_name = _cell_str(row, col_map.get("param_name"))
        yang_path = _cell_str(row, col_map.get("yang_path"))

        if not param_name or not yang_path:
            result.skipped_rows += 1
            result.warnings.append(f"행 {row_idx}: param_name 또는 yang_path 비어있음 — 스킵")
            continue

        result.records.append(ParameterRecord(
            param_name=param_name,
            yang_path=yang_path,
            description=_cell_str(row, col_map.get("description")),
            leaf_status=_cell_str(row, col_map.get("leaf_status")),
            units=_cell_str(row, col_map.get("units")),
            type=_cell_str(row, col_map.get("type")),
            pattern=_cell_str(row, col_map.get("pattern")),
            range=_cell_str(row, col_map.get("range")),
            default=_cell_str(row, col_map.get("default")),
            bandwidth_dependency=_cell_str(row, col_map.get("bandwidth_dependency")),
            config_value=_cell_str(row, col_map.get("config_value")),
            level=_cell_str(row, col_map.get("level")),
            restriction=_cell_str(row, col_map.get("restriction")),
            service_impact=_cell_str(row, col_map.get("service_impact")),
            realtime_change=_cell_str(row, col_map.get("realtime_change")),
            reference=_cell_str(row, col_map.get("reference")),
            mandatory=_cell_str(row, col_map.get("mandatory")),
            param_family=_cell_str(row, col_map.get("param_family")),
            related_feature_id=_cell_str(row, col_map.get("related_feature_id")),
            user_level=_cell_str(row, col_map.get("user_level")),
            feature_name=_cell_str(row, col_map.get("feature_name")),
            min=_cell_str(row, col_map.get("min")),
            max=_cell_str(row, col_map.get("max")),
        ))

    wb.close()
    return result
