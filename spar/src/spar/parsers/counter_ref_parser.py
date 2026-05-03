"""Samsung RAN 카운터 레퍼런스 Excel 파서.

문서 구조:
    대 그룹명(기술 계층) / 중 그룹명 / 중 그룹 ID — 병합 셀로 여러 카운터 행에 걸침.
    카운터명 / 설명 / 측정 주기 / 단위 / Min / Max — 각 행마다 존재.

병합 셀 처리: ws.merged_cells.ranges 기반 값 전파 (carry-forward fallback 병행).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

_COLUMN_ALIASES: dict[str, str] = {
    # 대 그룹
    "대분류": "large_group",
    "large group": "large_group",
    "layer": "large_group",
    "technology group": "large_group",
    "group category": "large_group",
    "tech group": "large_group",
    # 중 그룹명
    "중분류": "mid_group",
    "group name": "mid_group",
    "counter group": "mid_group",
    "category": "mid_group",
    "sub group": "mid_group",
    "subgroup": "mid_group",
    "중 그룹명": "mid_group",
    # 중 그룹 ID
    "id": "mid_group_id",
    "group id": "mid_group_id",
    "counter id": "mid_group_id",
    "no": "mid_group_id",
    "중 그룹 id": "mid_group_id",
    "mid group id": "mid_group_id",
    # 카운터명
    "counter name": "counter_name",
    "name": "counter_name",
    "카운터명": "counter_name",
    "counter": "counter_name",
    # 설명
    "description": "description",
    "desc": "description",
    "설명": "description",
    # 측정 주기
    "period": "period",
    "measurement period": "period",
    "주기": "period",
    "interval": "period",
    "collection period": "period",
    # 단위
    "unit": "unit",
    "단위": "unit",
    # 범위 (분리)
    "min": "min_val",
    "minimum": "min_val",
    "최솟값": "min_val",
    "min value": "min_val",
    "max": "max_val",
    "maximum": "max_val",
    "최댓값": "max_val",
    "max value": "max_val",
    # 범위 (통합)
    "value range": "value_range",
    "range": "value_range",
    "값 범위": "value_range",
    # 피처 ID
    "feature id": "feature_id",
    "feature": "feature_id",
    "피처 id": "feature_id",
    "feature_id": "feature_id",
}

REQUIRED_FIELDS = {"counter_name"}


def _parse_value_range(raw: str) -> tuple[str, str]:
    """'0 ~ 1024' 또는 '0-1024' → ('0', '1024')."""
    for sep in ("~", "-"):
        if sep in raw:
            parts = raw.split(sep, 1)
            return parts[0].strip(), parts[1].strip()
    return raw.strip(), ""


@dataclass
class CounterRecord:
    large_group: str
    mid_group: str
    mid_group_id: str
    counter_name: str
    description: str = ""
    period: str = ""
    unit: str = ""
    min_val: str = ""
    max_val: str = ""
    feature_id: str = ""

    def to_chunk_text(self) -> str:
        """RAG ingest용 텍스트. 그룹 계층 포함."""
        lines = [f"Counter: {self.counter_name}"]
        group_parts = [p for p in [self.large_group, self.mid_group] if p]
        group_str = " > ".join(group_parts)
        if group_str:
            id_suffix = f" [ID: {self.mid_group_id}]" if self.mid_group_id else ""
            lines.append(f"Group: {group_str}{id_suffix}")
        if self.description:
            lines.append(f"Description: {self.description}")
        if self.period:
            lines.append(f"Period: {self.period}")
        if self.unit:
            lines.append(f"Unit: {self.unit}")
        range_parts = []
        if self.min_val:
            range_parts.append(self.min_val)
        if self.max_val:
            range_parts.append(self.max_val)
        if range_parts:
            lines.append(f"Value Range: {' ~ '.join(range_parts)}")
        return "\n".join(lines)

    def to_dict(self) -> dict[str, Any]:
        return {
            "large_group": self.large_group,
            "mid_group": self.mid_group,
            "mid_group_id": self.mid_group_id,
            "counter_name": self.counter_name,
            "description": self.description,
            "period": self.period,
            "unit": self.unit,
            "min_val": self.min_val,
            "max_val": self.max_val,
        }


@dataclass
class CounterRefParseResult:
    records: list[CounterRecord] = field(default_factory=list)
    skipped_rows: int = 0
    warnings: list[str] = field(default_factory=list)


def _expand_merged_cells(ws) -> dict[tuple[int, int], Any]:
    """병합 범위 최상단 값을 모든 셀 좌표에 전파."""
    values: dict[tuple[int, int], Any] = {}
    for row in ws.iter_rows():
        for cell in row:
            values[(cell.row, cell.column)] = cell.value

    for merge_range in ws.merged_cells.ranges:
        top_val = values.get((merge_range.min_row, merge_range.min_col))
        for r in range(merge_range.min_row, merge_range.max_row + 1):
            for c in range(merge_range.min_col, merge_range.max_col + 1):
                values[(r, c)] = top_val

    return values


def _resolve_header(ws, expanded: dict[tuple[int, int], Any]) -> tuple[int, dict[str, int]]:
    """헤더 행 번호(1-based)와 {필드명: 열 인덱스(1-based)} 반환."""
    max_col = ws.max_column or 20
    for row_idx in range(1, min(11, (ws.max_row or 10) + 1)):
        col_map: dict[str, int] = {}
        for col_idx in range(1, max_col + 1):
            val = expanded.get((row_idx, col_idx))
            if val is None:
                continue
            alias = str(val).strip().lower()
            field_name = _COLUMN_ALIASES.get(alias)
            if field_name:
                col_map.setdefault(field_name, col_idx)
        if REQUIRED_FIELDS.issubset(col_map):
            return row_idx, col_map
    raise ValueError(
        "헤더 행 탐색 실패 — 필수 컬럼 없음: counter_name(또는 Counter Name, 카운터명). "
        "컬럼명 확인 필요."
    )


def _get(expanded: dict[tuple[int, int], Any], row: int, col: int | None) -> str:
    if col is None:
        return ""
    val = expanded.get((row, col))
    if val is None:
        return ""
    return str(val).strip()


def parse_counter_ref_excel(
    path: str | Path,
    sheet_name: str | None = None,
) -> CounterRefParseResult:
    """Excel 카운터 레퍼런스를 파싱해 CounterRefParseResult 반환."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"파일 없음: {path}")

    # read_only=False 필수 — merged_cells 정보 접근
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    expanded = _expand_merged_cells(ws)
    header_row, col_map = _resolve_header(ws, expanded)

    result = CounterRefParseResult()
    max_row = ws.max_row or 0

    for row_idx in range(header_row + 1, max_row + 1):
        counter_name = _get(expanded, row_idx, col_map.get("counter_name"))
        if not counter_name:
            result.skipped_rows += 1
            continue

        min_val = _get(expanded, row_idx, col_map.get("min_val"))
        max_val = _get(expanded, row_idx, col_map.get("max_val"))

        # value_range 단일 컬럼 처리
        if not min_val and not max_val:
            raw_range = _get(expanded, row_idx, col_map.get("value_range"))
            if raw_range:
                min_val, max_val = _parse_value_range(raw_range)

        result.records.append(CounterRecord(
            large_group=_get(expanded, row_idx, col_map.get("large_group")),
            mid_group=_get(expanded, row_idx, col_map.get("mid_group")),
            mid_group_id=_get(expanded, row_idx, col_map.get("mid_group_id")),
            counter_name=counter_name,
            description=_get(expanded, row_idx, col_map.get("description")),
            period=_get(expanded, row_idx, col_map.get("period")),
            unit=_get(expanded, row_idx, col_map.get("unit")),
            min_val=min_val,
            max_val=max_val,
            feature_id=_get(expanded, row_idx, col_map.get("feature_id")),
        ))

    wb.close()
    return result
