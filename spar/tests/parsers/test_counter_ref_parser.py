"""CounterRefParser 유닛 테스트."""
from __future__ import annotations

import pytest
import openpyxl
from pathlib import Path

from spar.parsers.counter_ref_parser import (
    CounterRecord,
    CounterRefParseResult,
    _parse_value_range,
    parse_counter_ref_excel,
)

SAMPLE = Path(__file__).parent.parent.parent / "data" / "samples" / "counter_ref_sample.xlsx"


# ── _parse_value_range ─────────────────────────────────────────────────────

class TestParseValueRange:
    def test_tilde_separator(self):
        assert _parse_value_range("0 ~ 1024") == ("0", "1024")

    def test_dash_separator(self):
        assert _parse_value_range("0-100") == ("0", "100")

    def test_no_separator(self):
        assert _parse_value_range("count") == ("count", "")

    def test_strips_whitespace(self):
        assert _parse_value_range("  -140  ~  -44  ") == ("-140", "-44")

    def test_negative_range_tilde(self):
        lo, hi = _parse_value_range("-140 ~ -44")
        assert lo == "-140" and hi == "-44"


# ── CounterRecord ──────────────────────────────────────────────────────────

class TestCounterRecord:
    def _make(self, **kwargs) -> CounterRecord:
        defaults = dict(
            large_group="RRC", mid_group="UE Statistics", mid_group_id="G-0042",
            counter_name="CELL.UE.MaxConnectedNbr",
            description="Max connected UEs", period="15 min", unit="count",
            min_val="0", max_val="1024",
        )
        defaults.update(kwargs)
        return CounterRecord(**defaults)

    def test_to_chunk_text_full(self):
        rec = self._make()
        text = rec.to_chunk_text()
        assert "Counter: CELL.UE.MaxConnectedNbr" in text
        assert "RRC > UE Statistics" in text
        assert "G-0042" in text
        assert "15 min" in text
        assert "count" in text
        assert "0 ~ 1024" in text

    def test_to_chunk_text_no_range(self):
        rec = self._make(min_val="", max_val="")
        text = rec.to_chunk_text()
        assert "Value Range" not in text

    def test_to_chunk_text_no_group(self):
        rec = self._make(large_group="", mid_group="", mid_group_id="")
        text = rec.to_chunk_text()
        assert "Group:" not in text

    def test_to_chunk_text_only_large_group(self):
        rec = self._make(mid_group="", mid_group_id="")
        text = rec.to_chunk_text()
        assert "Group: RRC" in text
        assert ">" not in text

    def test_to_dict_keys(self):
        rec = self._make()
        d = rec.to_dict()
        assert set(d.keys()) == {
            "large_group", "mid_group", "mid_group_id",
            "counter_name", "description", "period", "unit", "min_val", "max_val",
        }

    def test_to_dict_values(self):
        rec = self._make()
        d = rec.to_dict()
        assert d["counter_name"] == "CELL.UE.MaxConnectedNbr"
        assert d["mid_group_id"] == "G-0042"


# ── parse_counter_ref_excel — 샘플 파일 ────────────────────────────────────

@pytest.mark.skipif(not SAMPLE.exists(), reason="sample 파일 없음")
class TestParseCounterRefExcelSample:
    def setup_method(self):
        self.result = parse_counter_ref_excel(SAMPLE)

    def test_total_records(self):
        assert len(self.result.records) == 13

    def test_no_warnings_for_sample(self):
        # 샘플은 깨끗한 데이터여야 함
        assert self.result.warnings == []

    def test_large_group_propagated(self):
        rrc_records = [r for r in self.result.records if r.large_group == "RRC"]
        assert len(rrc_records) == 6

    def test_mac_large_group(self):
        mac_records = [r for r in self.result.records if r.large_group == "MAC"]
        assert len(mac_records) == 5  # PDSCH(3) + PUSCH(2)

    def test_phy_large_group(self):
        phy_records = [r for r in self.result.records if r.large_group == "PHY"]
        assert len(phy_records) == 2

    def test_mid_group_propagated(self):
        ue_stat = [r for r in self.result.records if r.mid_group == "UE Statistics"]
        assert len(ue_stat) == 3

    def test_mid_group_id_propagated(self):
        g0042 = [r for r in self.result.records if r.mid_group_id == "G-0042"]
        assert len(g0042) == 3

    def test_first_record_fields(self):
        rec = self.result.records[0]
        assert rec.counter_name == "CELL.UE.MaxConnectedNbr"
        assert rec.large_group == "RRC"
        assert rec.mid_group == "UE Statistics"
        assert rec.mid_group_id == "G-0042"
        assert rec.period == "15 min"
        assert rec.unit == "count"
        assert rec.min_val == "0"
        assert rec.max_val == "1024"

    def test_phy_rsrp_range(self):
        rsrp = next(r for r in self.result.records if r.counter_name == "CELL.PHY.AvgRSRP")
        assert rsrp.min_val == "-140"
        assert rsrp.max_val == "-44"

    def test_chunk_text_contains_group(self):
        rec = self.result.records[0]
        text = rec.to_chunk_text()
        assert "RRC > UE Statistics [ID: G-0042]" in text


# ── parse_counter_ref_excel — 합성 파일 ────────────────────────────────────

def _make_wb(headers: list[str], rows: list[tuple], merge_specs: list[str] | None = None, tmp_path=None):
    """테스트용 최소 Excel 생성."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)
    if merge_specs:
        for spec in merge_specs:
            ws.merge_cells(spec)
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


class TestParseCounterRefExcelSynthetic:
    def test_basic_no_merges(self, tmp_path):
        path = _make_wb(
            ["Large Group", "Group Name", "Group ID", "Counter Name", "Description", "Period", "Unit", "Min", "Max"],
            [("RRC", "UE Stats", "G-001", "CELL.UE.Count", "UE count", "15 min", "count", "0", "512")],
            tmp_path=tmp_path,
        )
        result = parse_counter_ref_excel(path)
        assert len(result.records) == 1
        assert result.records[0].counter_name == "CELL.UE.Count"
        assert result.records[0].large_group == "RRC"

    def test_merged_cell_propagation(self, tmp_path):
        rows = [
            ("RRC", "UE Stats", "G-001", "CELL.UE.A", "desc A", "15 min", "count", "0", ""),
            (None,  None,       None,    "CELL.UE.B", "desc B", "15 min", "count", "0", ""),
            (None,  None,       None,    "CELL.UE.C", "desc C", "15 min", "count", "0", ""),
        ]
        path = _make_wb(
            ["Large Group", "Group Name", "Group ID", "Counter Name", "Description", "Period", "Unit", "Min", "Max"],
            rows,
            merge_specs=["A2:A4", "B2:B4", "C2:C4"],
            tmp_path=tmp_path,
        )
        result = parse_counter_ref_excel(path)
        assert len(result.records) == 3
        for rec in result.records:
            assert rec.large_group == "RRC"
            assert rec.mid_group == "UE Stats"
            assert rec.mid_group_id == "G-001"

    def test_skips_empty_counter_name_rows(self, tmp_path):
        rows = [
            ("RRC", "UE Stats", "G-001", "CELL.UE.A", "desc", "15 min", "count", "0", ""),
            (None,  None,       None,    None,         "",    "",        "",      "",  ""),
            (None,  None,       None,    "CELL.UE.B", "desc", "15 min", "count", "0", ""),
        ]
        path = _make_wb(
            ["Large Group", "Group Name", "Group ID", "Counter Name", "Description", "Period", "Unit", "Min", "Max"],
            rows,
            tmp_path=tmp_path,
        )
        result = parse_counter_ref_excel(path)
        assert len(result.records) == 2
        assert result.skipped_rows == 1

    def test_value_range_single_column(self, tmp_path):
        path = _make_wb(
            ["Counter Name", "Description", "Value Range"],
            [("CELL.X", "desc", "0 ~ 100")],
            tmp_path=tmp_path,
        )
        result = parse_counter_ref_excel(path)
        assert result.records[0].min_val == "0"
        assert result.records[0].max_val == "100"

    def test_missing_required_column_raises(self, tmp_path):
        path = _make_wb(
            ["Description", "Period"],
            [("desc", "15 min")],
            tmp_path=tmp_path,
        )
        with pytest.raises(ValueError, match="헤더 행 탐색 실패"):
            parse_counter_ref_excel(path)

    def test_file_not_found_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            parse_counter_ref_excel(tmp_path / "nonexistent.xlsx")

    def test_korean_headers(self, tmp_path):
        path = _make_wb(
            ["대분류", "중분류", "중 그룹 ID", "카운터명", "설명", "주기", "단위", "최솟값", "최댓값"],
            [("RRC", "UE 통계", "G-001", "CELL.UE.Count", "UE 수", "15분", "개", "0", "1024")],
            tmp_path=tmp_path,
        )
        result = parse_counter_ref_excel(path)
        rec = result.records[0]
        assert rec.large_group == "RRC"
        assert rec.counter_name == "CELL.UE.Count"
        assert rec.min_val == "0"
        assert rec.max_val == "1024"

    def test_multiple_groups(self, tmp_path):
        rows = [
            ("RRC", "UE Stats",  "G-001", "CELL.UE.A", "", "15 min", "count", "", ""),
            (None,  "HO Stats",  "G-002", "CELL.HO.A", "", "15 min", "count", "", ""),
            ("MAC", "DL Stats",  "G-101", "CELL.DL.A", "", "15 min", "kbps",  "", ""),
        ]
        path = _make_wb(
            ["Large Group", "Group Name", "Group ID", "Counter Name", "Description", "Period", "Unit", "Min", "Max"],
            rows,
            tmp_path=tmp_path,
        )
        result = parse_counter_ref_excel(path)
        assert result.records[0].large_group == "RRC"
        assert result.records[0].mid_group == "UE Stats"
        assert result.records[1].mid_group == "HO Stats"
        assert result.records[2].large_group == "MAC"
