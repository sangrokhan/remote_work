from __future__ import annotations

import json
from pathlib import Path

import openpyxl


def load_excel_terms(path: str | Path, columns: list[str]) -> dict[str, dict]:
    """Excel 파일에서 지정 column 값을 추출해 {term: {"type": "keyword"}} 반환."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    col_indices = [headers.index(c) for c in columns if c in headers]

    terms: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        for idx in col_indices:
            if idx < len(row) and row[idx] is not None:
                term = str(row[idx]).strip()
                if term:
                    terms[term] = {"type": "keyword"}

    wb.close()
    return terms


def merge_into_acronyms(terms: dict[str, dict], acronyms_path: str | Path) -> None:
    """terms를 acronyms.json의 'keywords' 섹션에 병합. global/conflicts 섹션 보존."""
    path = Path(acronyms_path)
    if path.exists():
        data: dict = json.loads(path.read_text(encoding="utf-8"))
    else:
        data = {"global": {}, "conflicts": {}, "keywords": {}}

    if "keywords" not in data:
        data["keywords"] = {}

    for term, entry in terms.items():
        data["keywords"][term] = entry

    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
