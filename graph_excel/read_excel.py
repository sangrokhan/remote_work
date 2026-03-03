import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def _normalize_cell(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text if text else None


def _split_system_ids(value):
    if value is None:
        return []

    if isinstance(value, (list, tuple, set)):
        ids = []
        for item in value:
            ids.extend(_split_system_ids(item))
        return ids

    raw = str(value).strip()
    if not raw:
        return []

    parts = re.split(r"[;,|\n\r]+", raw)
    return [part.strip() for part in parts if part and part.strip()]


def _append_triple(triples, subject, predicate, object_value, metadata=None):
    triple = {
        "subject": subject,
        "predicate": predicate,
        "object": object_value,
    }
    if metadata:
        triple["meta"] = metadata
    triples.append(triple)


def process_counters_sheet(sheet, file_stem, triples):
    for row in range(2, sheet.max_row + 1):
        category = _normalize_cell(sheet[f"A{row}"].value)
        name = _normalize_cell(sheet[f"B{row}"].value)
        counter_id = _normalize_cell(sheet[f"C{row}"].value)
        system_ids = _split_system_ids(sheet[f"X{row}"].value)

        if not name or not counter_id:
            continue

        subject = name
        metadata = {
            "file": file_stem,
            "sheet": "Counter",
            "category": category,
            "row": row,
        }
        _append_triple(
            triples,
            subject=subject,
            predicate="HAS_ID",
            object_value=counter_id,
            metadata=metadata,
        )

        if category is not None:
            _append_triple(
                triples,
                subject=subject,
                predicate="HAS_CATEGORY",
                object_value=category,
                metadata=metadata,
            )

        for system_id in system_ids:
            _append_triple(
                triples,
                subject=subject,
                predicate="HAS_SYSTEM_ID",
                object_value=system_id,
                metadata=metadata,
            )


def process_parameters_sheet(sheet, file_stem, triples):
    for row in range(2, sheet.max_row + 1):
        parameter = _normalize_cell(sheet[f"B{row}"].value)
        system_ids = _split_system_ids(sheet[f"U{row}"].value)
        if not parameter or not system_ids:
            continue

        metadata = {
            "file": file_stem,
            "sheet": "Parameter Description",
            "row": row,
        }
        for system_id in system_ids:
            _append_triple(
                triples,
                subject=parameter,
                predicate="HAS_SYSTEM_ID",
                object_value=system_id,
                metadata=metadata,
            )


def preprocess_excel(path, file_type="auto"):
    workbook = load_workbook(filename=path, data_only=True, read_only=True)
    sheet_names = [name.lower() for name in workbook.sheetnames]
    file_stem = Path(path).stem
    triples = []

    if file_type in ("auto", "counters"):
        if "counter" in sheet_names:
            process_counters_sheet(workbook["Counter"], file_stem, triples)
        elif file_type == "counters":
            workbook.close()
            raise ValueError("Counter sheet not found. Expected sheet name: 'Counter'.")

    if file_type in ("auto", "parameters"):
        if "parameter description" in sheet_names:
            process_parameters_sheet(
                workbook["Parameter Description"],
                file_stem,
                triples,
            )
        elif file_type == "parameters":
            workbook.close()
            raise ValueError(
                "Parameter sheet not found. Expected sheet name: 'Parameter Description'."
            )

    workbook.close()

    if file_type == "auto" and not triples:
        raise ValueError(
            "No supported sheet found. Expected 'Counter' or 'Parameter Description'."
        )
    return triples


def write_triples(triples, output_path):
    with open(output_path, "w", encoding="utf-8") as f:
        for triple in triples:
            f.write(json.dumps(triple, ensure_ascii=False))
            f.write("\n")


def parse_args():
    parser = argparse.ArgumentParser(description="Preprocess Excel files into Neo4j triple JSON.")
    parser.add_argument("file", help="Path to the Excel file")
    parser.add_argument(
        "--type",
        choices=["auto", "counters", "parameters"],
        default="auto",
        help="Target sheet type. Auto detects based on sheet names.",
    )
    parser.add_argument(
        "--output",
        help="Output path for triples (.jsonl recommended). Defaults to <filename>_triples.jsonl",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    output_path = args.output or f"{Path(args.file).stem}_triples.jsonl"
    triples = preprocess_excel(args.file, file_type=args.type)
    write_triples(triples, output_path)
    print(f"Generated {len(triples)} triples -> {output_path}")


if __name__ == "__main__":
    main()
