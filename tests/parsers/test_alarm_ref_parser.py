from pathlib import Path

import openpyxl
import pytest

from spar.parsers.alarm_ref_parser import (
    AlarmRecord,
    parse_alarm_ref_excel,
    _classify_alarm_code,
)


def _write_xlsx_single(tmp_path: Path, rows: list[tuple], sheet_name: str = "Sheet") -> Path:
    """단일 시트 Excel 생성."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r in rows:
        ws.append(r)
    p = tmp_path / "t.xlsx"
    wb.save(p)
    return p


def _write_xlsx_multi(tmp_path: Path, sheets: dict[str, list[tuple]]) -> Path:
    """다중 시트 Excel 생성. sheets = {sheet_name: rows}"""
    wb = openpyxl.Workbook()
    first = True
    for name, rows in sheets.items():
        if first:
            ws = wb.active
            ws.title = name
            first = False
        else:
            ws = wb.create_sheet(name)
        for r in rows:
            ws.append(r)
    p = tmp_path / "multi.xlsx"
    wb.save(p)
    return p


# ── 알람 코드 분류 ──────────────────────────────────────────────

def test_classify_ran_code():
    assert _classify_alarm_code("A1234567R") == "RAN"
    assert _classify_alarm_code("A0000000R") == "RAN"


def test_classify_non_ran_code():
    assert _classify_alarm_code("A123D") == "NON_RAN"
    assert _classify_alarm_code("H456D") == "NON_RAN"


def test_classify_unknown_code():
    assert _classify_alarm_code("ALM-1001") == "UNKNOWN"
    assert _classify_alarm_code("B123D") == "UNKNOWN"
    assert _classify_alarm_code("A12345D") == "UNKNOWN"


# ── 다중 시트 파싱 (가이드 시트 스킵) ─────────────────────────

def test_multi_sheet_skips_first_guide_sheet(tmp_path):
    p = _write_xlsx_multi(tmp_path, {
        "Guide": [
            ("This is a guide sheet", None, None, None),
        ],
        "Alarm List": [
            ("Alarm Code", "Alarm Name", "Severity", "Group"),
            ("A1234567R", "RRU Fault", "Critical", "Radio"),
            ("A123D", "HW Error", "Major", "HW"),
        ],
    })
    res = parse_alarm_ref_excel(p)
    assert len(res.records) == 2
    assert res.records[0].alarm_id == "A1234567R"
    assert res.records[0].alarm_type == "RAN"
    assert res.records[1].alarm_id == "A123D"
    assert res.records[1].alarm_type == "NON_RAN"


def test_multi_sheet_parses_all_data_sheets(tmp_path):
    p = _write_xlsx_multi(tmp_path, {
        "Guide": [("guide",)],
        "RAN Alarms": [
            ("Alarm Code", "Alarm Name", "Severity", "Group"),
            ("A1111111R", "BBU Fault", "Critical", "BBU"),
        ],
        "Non-RAN Alarms": [
            ("Alarm Code", "Alarm Name", "Severity", "Group"),
            ("H001D", "NTP Loss", "Major", "Sync"),
        ],
    })
    res = parse_alarm_ref_excel(p)
    assert len(res.records) == 2
    ids = {r.alarm_id for r in res.records}
    assert ids == {"A1111111R", "H001D"}


def test_single_sheet_fallback_with_warning(tmp_path):
    """시트 1개뿐이면 경고 후 파싱."""
    p = _write_xlsx_single(tmp_path, [
        ("Alarm Code", "Alarm Name", "Severity", "Group"),
        ("A9999999R", "Test", "Minor", "Core"),
    ])
    res = parse_alarm_ref_excel(p)
    assert len(res.records) == 1
    assert any("1개뿐" in w for w in res.warnings)


# ── 특정 시트 지정 파싱 ────────────────────────────────────────

def test_sheet_name_param(tmp_path):
    p = _write_xlsx_multi(tmp_path, {
        "Guide": [("guide",)],
        "Alarm List": [
            ("Alarm Code", "Alarm Name", "Severity", "Group"),
            ("A2345678R", "Cell Down", "Critical", "Radio"),
        ],
    })
    res = parse_alarm_ref_excel(p, sheet_name="Alarm List")
    assert len(res.records) == 1
    assert res.records[0].alarm_id == "A2345678R"


# ── 헤더 별칭 해석 ─────────────────────────────────────────────

def test_header_alias_resolution(tmp_path):
    p = _write_xlsx_single(tmp_path, [
        ("Code", "Name", "Level", "Category"),
        ("A1234567R", "Test", "Major", "RAN"),
    ])
    res = parse_alarm_ref_excel(p)
    assert len(res.records) == 1
    assert res.records[0].alarm_id == "A1234567R"
    assert res.records[0].group == "RAN"


# ── 필수 필드 누락 스킵 ────────────────────────────────────────

def test_required_field_missing_skipped(tmp_path):
    p = _write_xlsx_single(tmp_path, [
        ("Alarm Code", "Alarm Name", "Severity", "Group"),
        ("", "Orphan", "Critical", ""),
        ("A1234567R", "", "Critical", ""),
        ("A123D", "Good", "Major", "HW"),
    ])
    res = parse_alarm_ref_excel(p)
    assert len(res.records) == 1
    assert res.records[0].alarm_id == "A123D"
    assert res.skipped_rows == 2


# ── 미지정 알람 코드 경고 ──────────────────────────────────────

def test_unknown_alarm_code_warns(tmp_path):
    p = _write_xlsx_single(tmp_path, [
        ("Alarm Code", "Alarm Name", "Severity", "Group"),
        ("BADCODE", "Unknown Alarm", "Minor", ""),
    ])
    res = parse_alarm_ref_excel(p)
    assert len(res.records) == 1
    assert res.records[0].alarm_type == "UNKNOWN"
    assert any("BADCODE" in w for w in res.warnings)


# ── AlarmRecord 메서드 ─────────────────────────────────────────

def test_to_keywords_excludes_blanks():
    rec = AlarmRecord(alarm_id="A123D", alarm_name="X")
    kw = rec.to_keywords()
    assert "A123D" in kw
    assert "X" in kw
    assert "NON_RAN" in kw
    assert "" not in kw


def test_to_chunk_text_format():
    rec = AlarmRecord(
        alarm_id="A1234567R",
        alarm_name="Cell Down",
        severity="Critical",
        group="Radio",
    )
    txt = rec.to_chunk_text()
    assert "Alarm: A1234567R — Cell Down" in txt
    assert "Type: RAN" in txt
    assert "Severity: Critical" in txt
    assert "Group: Radio" in txt


def test_to_dict_fields():
    rec = AlarmRecord(alarm_id="H001D", alarm_name="NTP Loss", group="Sync")
    d = rec.to_dict()
    assert d["alarm_id"] == "H001D"
    assert d["alarm_type"] == "NON_RAN"
    assert d["group"] == "Sync"
    assert "category" not in d
    assert "module" not in d
