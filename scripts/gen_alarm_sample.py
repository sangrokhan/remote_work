"""Reproducible generator for alarm_excel_ref_sample.xlsx.

Run: python scripts/gen_alarm_sample.py
Outputs: data/samples/alarm_excel_ref_sample.xlsx

Sheet layout:
    Sheet 0 — Guide    (skipped by parser)
    Sheet 1 — RAN Alarms
    Sheet 2 — Non-RAN Alarms
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

HEADERS = ("Alarm Code", "Alarm Name", "Severity", "Group")

# RAN: A + 7 digits + R  (ne=001, unit=00, alarm_type=1..8)
RAN_ROWS: list[tuple[str, str, str, str]] = [
    ("A0010001R", "Cell Out of Service",    "Critical", "Radio"),
    ("A0010002R", "Link Down",              "Critical", "Transport"),
    ("A0010003R", "Cell Down",              "Critical", "Radio"),
    ("A0010004R", "High Temperature",       "Major",    "HW"),
    ("A0010005R", "Fan Failure",            "Major",    "HW"),
    ("A0010006R", "Clock Sync Loss",        "Major",    "Transport"),
    ("A0010007R", "PRACH Anomaly",          "Minor",    "Radio"),
    ("A0010008R", "Power Redundancy Lost",  "Major",    "HW"),
]

# Non-RAN: [AH] + 3 digits + D
NON_RAN_ROWS: list[tuple[str, str, str, str]] = [
    ("A001D", "License Expiring",   "Minor",   "SW"),
    ("A002D", "High CPU Usage",     "Minor",   "SW"),
    ("A003D", "Backup Failed",      "Warning", "SW"),
    ("A004D", "Config Mismatch",    "Warning", "SW"),
    ("H001D", "NTP Sync Loss",      "Major",   "Sync"),
    ("H002D", "Disk Full",          "Major",   "HW"),
    ("H003D", "Auth Failure",       "Critical","Security"),
    ("H004D", "Certificate Expiry", "Warning", "Security"),
]

GUIDE_TEXT = [
    ("Samsung RAN Alarm Reference",),
    ("",),
    ("Sheet structure:",),
    ("  Sheet 1 — RAN Alarms     : code format A + 7 digits + R",),
    ("  Sheet 2 — Non-RAN Alarms : code format [AH] + 3 digits + D",),
    ("",),
    ("Column order: Alarm Code | Alarm Name | Severity | Group",),
]


def _append_data_sheet(wb: openpyxl.Workbook, title: str, rows: list[tuple]) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(HEADERS)
    # Bold header
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)


def main() -> None:
    out = Path(__file__).resolve().parents[1] / "data" / "samples" / "alarm_excel_ref_sample.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # Sheet 0: Guide
    ws_guide = wb.active
    ws_guide.title = "Guide"
    for row in GUIDE_TEXT:
        ws_guide.append(row)

    # Sheet 1 & 2: Alarm data
    _append_data_sheet(wb, "RAN Alarms",     RAN_ROWS)
    _append_data_sheet(wb, "Non-RAN Alarms", NON_RAN_ROWS)

    wb.save(out)
    total = len(RAN_ROWS) + len(NON_RAN_ROWS)
    print(f"wrote {out} ({total} rows across 2 data sheets)")


if __name__ == "__main__":
    main()
