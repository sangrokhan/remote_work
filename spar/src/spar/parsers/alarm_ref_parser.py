"""Samsung RAN 알람 레퍼런스 Excel 파서.

지원 컬럼 (헤더명 자동 탐색, 별칭 허용):
    Alarm ID, Alarm Name, Severity, Category, Module
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

_COLUMN_ALIASES: dict[str, str] = {
    "alarm id": "alarm_id",
    "id": "alarm_id",
    "alarm code": "alarm_id",
    "code": "alarm_id",
    "alarm name": "alarm_name",
    "name": "alarm_name",
    "severity": "severity",
    "level": "severity",
    "category": "category",
    "group": "category",
    "module": "module",
    "node": "module",
    "subsystem": "module",
}

REQUIRED_FIELDS = {"alarm_id", "alarm_name"}


def _normalize_alarm_id(raw: str) -> str:
    return raw.strip().upper()


@dataclass
class AlarmRecord:
    alarm_id: str
    alarm_name: str
    severity: str = ""
    category: str = ""
    module: str = ""
    pdf_ref: str = ""

    def __post_init__(self) -> None:
        self.alarm_id = _normalize_alarm_id(self.alarm_id)

    def to_chunk_text(self) -> str:
        lines = [f"Alarm: {self.alarm_id} — {self.alarm_name}"]
        if self.severity:
            lines.append(f"Severity: {self.severity}")
        if self.category:
            lines.append(f"Category: {self.category}")
        if self.module:
            lines.append(f"Module: {self.module}")
        return "\n".join(lines)

    def to_keywords(self) -> list[str]:
        return [v for v in (self.alarm_id, self.alarm_name, self.severity,
                            self.category, self.module) if v]

    def to_dict(self) -> dict[str, Any]:
        return {
            "alarm_id": self.alarm_id,
            "alarm_name": self.alarm_name,
            "severity": self.severity,
            "category": self.category,
            "module": self.module,
            "pdf_ref": self.pdf_ref,
        }


@dataclass
class AlarmRefParseResult:
    records: list[AlarmRecord] = field(default_factory=list)
    skipped_rows: int = 0
    warnings: list[str] = field(default_factory=list)


def _resolve_header_row(ws) -> tuple[int, dict[str, int]]:
    for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), start=1):
        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            alias = str(cell).strip().lower()
            field_name = _COLUMN_ALIASES.get(alias)
            if field_name and field_name not in col_map:
                col_map[field_name] = col_idx
        if REQUIRED_FIELDS.issubset(col_map):
            return row_idx, col_map
    raise ValueError(
        f"헤더 행 탐색 실패 — 필수 컬럼 없음: {REQUIRED_FIELDS}. "
        "컬럼명 확인 필요 (Alarm ID, Alarm Name)"
    )


def _cell_str(row: tuple, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    val = row[idx]
    if val is None:
        return ""
    return str(val).strip()


def parse_alarm_ref_excel(
    path: str | Path,
    sheet_name: str | None = None,
) -> AlarmRefParseResult:
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"파일 없음: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row, col_map = _resolve_header_row(ws)

    result = AlarmRefParseResult()
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        if all(v is None for v in row):
            continue

        alarm_id = _cell_str(row, col_map.get("alarm_id"))
        alarm_name = _cell_str(row, col_map.get("alarm_name"))

        if not alarm_id or not alarm_name:
            result.skipped_rows += 1
            result.warnings.append(
                f"행 {row_idx}: alarm_id 또는 alarm_name 비어있음 — 스킵"
            )
            continue

        result.records.append(
            AlarmRecord(
                alarm_id=alarm_id,
                alarm_name=alarm_name,
                severity=_cell_str(row, col_map.get("severity")),
                category=_cell_str(row, col_map.get("category")),
                module=_cell_str(row, col_map.get("module")),
            )
        )

    wb.close()
    return result
